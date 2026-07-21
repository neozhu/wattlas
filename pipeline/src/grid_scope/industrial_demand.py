from __future__ import annotations

from collections.abc import Iterable, Mapping
from datetime import date, datetime
import json
from pathlib import Path
import math
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook

from grid_scope.canonicalize import assign_asset_geography, build_geography_index


HYDROGEN_PRODUCTION_SOURCE_ID = "iea-hydrogen-production-2026"
HYDROGEN_INFRASTRUCTURE_SOURCE_ID = "iea-hydrogen-infrastructure-2026"
GEM_CEMENT_SOURCE_ID = "gem-global-cement-concrete-2025"
GEM_STEEL_PLANT_SOURCE_ID = "gem-global-steel-plants-2026"
GEM_STEEL_UNIT_SOURCE_ID = "gem-global-steel-units-2026"
GEM_IRON_UNIT_SOURCE_ID = "gem-global-iron-units-2026"


def _key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "_", text.casefold()).strip("_")


def _slug(value: object) -> str:
    return _key(value).replace("_", "-") or "unknown"


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.casefold() in {"n/a", "na", "none", "unknown", "tbc", "-"}:
        return None
    return text


def _number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).replace(",", "").strip()
        match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if match is None:
            return None
        number = float(match.group())
    if not math.isfinite(number) or number < 0:
        return None
    return number


def _capacity_mwel(value: object) -> float | None:
    number = _number(value)
    if number is None:
        return None
    text = str(value).casefold()
    if "gw" in text and "gwh" not in text:
        number *= 1_000
    elif "kw" in text and "kwh" not in text:
        number /= 1_000
    return round(number, 6)


def _year(value: object) -> int | None:
    if isinstance(value, (date, datetime)):
        return value.year
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    match = re.search(r"(?:19|20)\d{2}", str(value or ""))
    if match is None:
        return None
    return int(match.group())


def _country(value: object, mapping: Mapping[str, str]) -> str | None:
    code = (_text(value) or "").upper()
    if len(code) == 2:
        return code
    return mapping.get(code)


def _coordinate(value: object, *, latitude: bool) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value).replace(",", "."))
        if match is None:
            return None
        number = float(match.group())
    if not math.isfinite(number):
        return None
    maximum = 90 if latitude else 180
    return number if -maximum <= number <= maximum else None


def _coordinate_pair(value: object) -> list[float] | None:
    if value is None:
        return None
    parts = [part.strip() for part in str(value).split(",")]
    if len(parts) != 2:
        return None
    latitude = _coordinate(parts[0], latitude=True)
    longitude = _coordinate(parts[1], latitude=False)
    if latitude is None or longitude is None:
        return None
    return [longitude, latitude]


def _country_name(value: object, mapping: Mapping[str, str]) -> str | None:
    name = _key(value)
    return mapping.get(name)


def _lifecycle(value: object) -> str:
    status = _key(value)
    if not status:
        return "unknown"
    if any(token in status for token in ("decomm", "retired")):
        return "decommissioned"
    if any(token in status for token in ("cancel", "removed")):
        return "cancelled"
    if any(token in status for token in ("dormant", "on_hold", "paused", "mothball", "shelved")):
        return "paused"
    if any(token in status for token in ("operational", "operating")):
        return "operational"
    if any(token in status for token in ("fid", "construction", "under_construction")):
        return "under_construction"
    if any(token in status for token in ("feasibility", "feed", "permit", "planning", "pre_construction")):
        return "pre_construction"
    if any(token in status for token in ("concept", "announced", "demo")):
        return "announced"
    return "unknown"


def _grid_connection(value: object) -> str:
    electricity = _key(value)
    if electricity in {"grid_renewables", "grid_plus_renewables"} or (
        "grid" in electricity and "renew" in electricity
    ):
        return "grid_plus_renewables"
    if "dedicated" in electricity and "renew" in electricity:
        return "dedicated_renewable"
    if "nuclear" in electricity:
        return "nuclear"
    if electricity == "grid" or electricity.startswith("grid_"):
        return "grid"
    return "other_or_unknown"


def _is_electrolysis(technology: object) -> bool:
    text = _key(technology)
    return any(token in text for token in ("electrolysis", "pem", "alk", "soec", "aem"))


def _records(path: Path, sheet_name: str, *, header_row: int) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header_values = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        headers = [_key(value) if value is not None else "" for value in header_values]
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            record = {
                header: value
                for header, value in zip(headers, values, strict=False)
                if header and value is not None
            }
            if record:
                yield record
    finally:
        workbook.close()


def _pick(record: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in record:
            return record[key]
    return None


def parse_hydrogen_production(
    path: Path, *, iso3_to_iso2: Mapping[str, str]
) -> list[dict[str, Any]]:
    assets: list[dict[str, Any]] = []
    for row in _records(path, "Hydrogen production projects", header_row=2):
        reference = _text(_pick(row, "reference"))
        name = _text(_pick(row, "project_name"))
        country = _country(_pick(row, "country_iso_3"), iso3_to_iso2)
        if reference is None or name is None or country is None:
            continue
        capacity = _capacity_mwel(_pick(row, "capacity_mwel"))
        grid_connection = _grid_connection(
            _pick(row, "type_of_electricity_for_electrolysis_projects")
        )
        technology = _text(_pick(row, "technology"))
        latitude = _coordinate(_pick(row, "latitude"), latitude=True)
        longitude = _coordinate(_pick(row, "longitude"), latitude=False)
        year = _year(_pick(row, "date_online"))
        status = _text(_pick(row, "status"))
        assets.append(
            {
                "id": f"iea-h2-production-{_slug(reference)}",
                "name": name,
                "geographyId": "",
                "country": country,
                "category": "industrial_load",
                "subtype": "hydrogen_production",
                "lifecycle": _lifecycle(status),
                "rawStatus": status,
                "targetYear": year if year is not None and 2026 <= year <= 2031 else None,
                "coordinates": [longitude, latitude]
                if longitude is not None and latitude is not None
                else None,
                "locationName": _text(_pick(row, "location")),
                "locationPrecision": "exact"
                if longitude is not None and latitude is not None
                else "city_centroid",
                "reportedCapacity": capacity,
                "reportedCapacityUnit": "MWel" if capacity is not None else None,
                "gridConnectionType": grid_connection,
                "gridDemandEligible": bool(
                    capacity is not None
                    and grid_connection in {"grid", "grid_plus_renewables"}
                    and _is_electrolysis(technology)
                ),
                "gridDemandContribution": False,
                "annualDemandGwh": None,
                "technologyDetail": _text(_pick(row, "technology_details")) or technology,
                "operator": None,
                "demandMw": None,
                "valueKind": "reported",
                "sourceType": "research_verified",
                "sourceIds": [HYDROGEN_PRODUCTION_SOURCE_ID],
                "sourceRecordIds": [reference],
                "externalIds": {"ieaHydrogenReference": reference},
                "confidence": 84 if capacity is not None else 72,
            }
        )
    return sorted(assets, key=lambda asset: asset["id"])


def _network_subtype(sheet_name: str, row: Mapping[str, Any]) -> str:
    if sheet_name == "H2 TRANSMISSION (pipe)":
        return "hydrogen_pipeline"
    if sheet_name == "H2 BLENDING":
        return "hydrogen_blending"
    if sheet_name == "UNDERGROUND H2 STORAGE":
        return "hydrogen_storage"
    trade = _key(_pick(row, "trade", "import_export_bunkering"))
    if "export" in trade and "import" not in trade:
        return "hydrogen_export_terminal"
    return "hydrogen_import_terminal"


def parse_hydrogen_infrastructure(
    path: Path, *, iso3_to_iso2: Mapping[str, str]
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = set(workbook.sheetnames)
    workbook.close()
    supported = (
        "H2 TRANSMISSION (pipe)",
        "H2 BLENDING",
        "UNDERGROUND H2 STORAGE",
        "NH3 INFRASTRUCTURE AT PORTS",
        "NH3 CRACKING PLANTS",
        "H2 INFRASTRUCTURE AT PORTS",
        "MeOH INFRASTRUCTURE AT PORTS",
    )
    assets: list[dict[str, Any]] = []
    for sheet_name in supported:
        if sheet_name not in sheet_names:
            continue
        for row in _records(path, sheet_name, header_row=1):
            reference = _text(_pick(row, "ref", "reference"))
            name = _text(_pick(row, "project_name"))
            country = _country(
                _pick(row, "country_1_iso", "country_iso"), iso3_to_iso2
            )
            if reference is None or name is None or country is None:
                continue
            secondary_country = _country(_pick(row, "country_2_iso"), iso3_to_iso2)
            status = _text(_pick(row, "status", "availability"))
            year = _year(_pick(row, "date_online", "announced_start_date"))
            capacity = _capacity_mwel(
                _pick(row, "capacity_mwel", "injection_capacity_mwel")
            )
            location = _text(_pick(row, "location", "port"))
            assets.append(
                {
                    "id": f"iea-h2-network-{_slug(reference)}",
                    "name": name,
                    "geographyId": "",
                    "country": country,
                    "secondaryCountries": [secondary_country]
                    if secondary_country and secondary_country != country
                    else [],
                    "category": "hydrogen_infrastructure",
                    "subtype": _network_subtype(sheet_name, row),
                    "lifecycle": _lifecycle(status),
                    "rawStatus": status,
                    "targetYear": year if year is not None and 2026 <= year <= 2031 else None,
                    "coordinates": None,
                    "locationName": location,
                    "locationPrecision": "city_centroid",
                    "reportedCapacity": capacity,
                    "reportedCapacityUnit": "MWel" if capacity is not None else None,
                    "gridConnectionType": None,
                    "gridDemandEligible": False,
                    "gridDemandContribution": False,
                    "annualDemandGwh": None,
                    "technologyDetail": _text(
                        _pick(row, "technology", "pipeline_type", "repurposed_new")
                    ),
                    "operator": _text(_pick(row, "partners", "participants")),
                    "demandMw": None,
                    "valueKind": "reported",
                    "sourceType": "research_verified",
                    "sourceIds": [HYDROGEN_INFRASTRUCTURE_SOURCE_ID],
                    "sourceRecordIds": [reference],
                    "externalIds": {"ieaHydrogenInfrastructureReference": reference},
                    "confidence": 78 if location else 68,
                }
            )
    return sorted(assets, key=lambda asset: asset["id"])


def parse_gem_cement(
    path: Path, *, country_name_to_iso2: Mapping[str, str]
) -> list[dict[str, Any]]:
    assets: list[dict[str, Any]] = []
    for row in _records(path, "Plant Data", header_row=1):
        plant_id = _text(_pick(row, "gem_plant_id"))
        name = _text(_pick(row, "gem_asset_name_english"))
        country = _country_name(_pick(row, "country_area"), country_name_to_iso2)
        coordinates = _coordinate_pair(_pick(row, "coordinates"))
        if plant_id is None or name is None or country is None or coordinates is None:
            continue
        status = _text(_pick(row, "operating_status"))
        year = _year(_pick(row, "start_date"))
        cement_capacity = _number(
            _pick(row, "cement_capacity_millions_metric_tonnes_per_annum")
        )
        clinker_capacity = _number(
            _pick(row, "clinker_capacity_millions_metric_tonnes_per_annum")
        )
        reported_capacity = cement_capacity if cement_capacity is not None else clinker_capacity
        assets.append(
            {
                "id": f"gem-cement-{_slug(plant_id)}",
                "name": name,
                "geographyId": "",
                "country": country,
                "category": "industrial_load",
                "subtype": "cement_plant",
                "lifecycle": _lifecycle(status),
                "rawStatus": status,
                "targetYear": year if year is not None and 2026 <= year <= 2031 else None,
                "coordinates": coordinates,
                "locationName": _text(_pick(row, "municipality")),
                "locationPrecision": "exact"
                if _key(_pick(row, "coordinate_accuracy")) == "exact"
                else "city_centroid",
                "cementCapacityMtpa": cement_capacity,
                "clinkerCapacityMtpa": clinker_capacity,
                "reportedCapacity": reported_capacity,
                "reportedCapacityUnit": "Mtpa" if reported_capacity is not None else None,
                "gridConnectionType": "other_or_unknown",
                "gridDemandEligible": bool(cement_capacity is not None),
                "gridDemandContribution": False,
                "annualDemandGwh": None,
                "technologyDetail": " · ".join(
                    value
                    for value in (
                        _text(_pick(row, "plant_type")),
                        _text(_pick(row, "production_type")),
                        _text(_pick(row, "majority_cement_type")),
                    )
                    if value
                ) or None,
                "operator": _text(_pick(row, "owner_name_english")),
                "demandMw": None,
                "valueKind": "reported",
                "sourceType": "research_verified",
                "sourceIds": [GEM_CEMENT_SOURCE_ID],
                "sourceRecordIds": [plant_id],
                "externalIds": {"gemPlantId": plant_id},
                "projectUrl": _text(_pick(row, "gem_wiki_page")),
                "confidence": 86 if _key(_pick(row, "coordinate_accuracy")) == "exact" else 76,
            }
        )
    return sorted(assets, key=lambda asset: asset["id"])


def _unit_index(path: Path, sheet_name: str) -> dict[str, list[dict[str, Any]]]:
    indexed: dict[str, list[dict[str, Any]]] = {}
    for row in _records(path, sheet_name, header_row=1):
        plant_id = _text(_pick(row, "gem_plant_id"))
        unit_id = _text(_pick(row, "gem_unit_id"))
        if plant_id is None or unit_id is None:
            continue
        indexed.setdefault(plant_id, []).append(
            {
                "unitId": unit_id,
                "status": _text(_pick(row, "unit_status")),
                "year": _year(
                    _pick(row, "start_date", "construction_date", "announced_date")
                ),
                "capacityTtpa": _number(_pick(row, "current_capacity_ttpa")),
                "projectUrl": _text(_pick(row, "gem_wiki_page")),
                "furnaceType": _text(_pick(row, "furnace_type")),
                "reductant": _text(_pick(row, "current_or_initial_reductant")),
                "hydrogenStatus": _text(_pick(row, "hydrogen_reductant_status")),
            }
        )
    return indexed


def parse_gem_steel(
    plant_path: Path,
    *,
    steel_units_path: Path,
    iron_units_path: Path,
    country_name_to_iso2: Mapping[str, str],
) -> list[dict[str, Any]]:
    plants: dict[str, dict[str, Any]] = {}
    for row in _records(plant_path, "Plant data", header_row=1):
        plant_id = _text(_pick(row, "gem_plant_id"))
        if plant_id is None:
            continue
        plants[plant_id] = {
            "name": _text(_pick(row, "plant_name_english")),
            "country": _country_name(_pick(row, "country_area"), country_name_to_iso2),
            "coordinates": _coordinate_pair(_pick(row, "coordinates")),
            "coordinateAccuracy": _key(_pick(row, "coordinate_accuracy")),
            "locationName": _text(_pick(row, "municipality")),
            "operator": _text(_pick(row, "owner")),
            "projectUrl": _text(_pick(row, "gem_wiki_page")),
            "powerSource": _text(_pick(row, "power_source")),
        }
    eaf_units = _unit_index(steel_units_path, "Electric arc furnaces")
    dri_units = _unit_index(iron_units_path, "Direct reduced iron furnaces")
    assets: list[dict[str, Any]] = []
    for row_number, row in enumerate(
        _records(plant_path, "Plant capacities and status", header_row=1), start=2
    ):
        plant_id = _text(_pick(row, "gem_plant_id"))
        plant = plants.get(plant_id or "")
        if plant_id is None or plant is None:
            continue
        if plant["name"] is None or plant["country"] is None or plant["coordinates"] is None:
            continue
        status = _text(_pick(row, "status"))
        year = _year(_pick(row, "start_date"))
        lifecycle = _lifecycle(status)
        eaf_capacity = _number(_pick(row, "nominal_eaf_steel_capacity_ttpa"))
        dri_capacity = _number(_pick(row, "nominal_dri_capacity_ttpa"))
        induction_capacity = _number(_pick(row, "nominal_if_steel_capacity_ttpa"))
        related_eaf = eaf_units.get(plant_id, [])
        related_dri = dri_units.get(plant_id, [])
        if eaf_capacity is None:
            values = [unit["capacityTtpa"] for unit in related_eaf if unit["capacityTtpa"] is not None]
            eaf_capacity = sum(values) if values else None
        if dri_capacity is None:
            values = [unit["capacityTtpa"] for unit in related_dri if unit["capacityTtpa"] is not None]
            dri_capacity = sum(values) if values else None
        unit_ids = sorted(
            {unit["unitId"] for unit in [*related_eaf, *related_dri]}
        )
        crude_capacity = _number(_pick(row, "nominal_crude_steel_capacity_ttpa"))
        equipment = _text(_pick(row, "main_production_equipment"))
        id_suffix = f"{_slug(status)}-{year or row_number}"
        assets.append(
            {
                "id": f"gem-steel-{_slug(plant_id)}-{id_suffix}",
                "name": plant["name"],
                "geographyId": "",
                "country": plant["country"],
                "category": "industrial_load",
                "subtype": "steel_plant",
                "lifecycle": lifecycle,
                "rawStatus": status,
                "targetYear": year if year is not None and 2026 <= year <= 2031 else None,
                "coordinates": plant["coordinates"],
                "locationName": plant["locationName"],
                "locationPrecision": "exact"
                if plant["coordinateAccuracy"] == "exact"
                else "city_centroid",
                "electricArcCapacityTtpa": eaf_capacity,
                "driCapacityTtpa": dri_capacity,
                "inductionCapacityTtpa": induction_capacity,
                "reportedCapacity": crude_capacity,
                "reportedCapacityUnit": "ktpa" if crude_capacity is not None else None,
                "gridConnectionType": "grid"
                if "grid" in _key(plant["powerSource"])
                else "other_or_unknown",
                "gridDemandEligible": bool(eaf_capacity is not None or induction_capacity is not None),
                "gridDemandContribution": False,
                "annualDemandGwh": None,
                "technologyDetail": equipment,
                "operator": plant["operator"],
                "demandMw": None,
                "valueKind": "reported",
                "sourceType": "research_verified",
                "sourceIds": sorted(
                    {
                        GEM_STEEL_PLANT_SOURCE_ID,
                        *([GEM_STEEL_UNIT_SOURCE_ID] if related_eaf else []),
                        *([GEM_IRON_UNIT_SOURCE_ID] if related_dri else []),
                    }
                ),
                "sourceRecordIds": sorted({plant_id, *unit_ids}),
                "externalIds": {"gemPlantId": plant_id},
                "projectUrl": plant["projectUrl"]
                or next(
                    (
                        unit["projectUrl"]
                        for unit in [*related_eaf, *related_dri]
                        if unit["projectUrl"]
                    ),
                    None,
                ),
                "confidence": 86 if plant["coordinateAccuracy"] == "exact" else 76,
                "driHydrogenContext": [
                    {
                        "unitId": unit["unitId"],
                        "furnaceType": unit["furnaceType"],
                        "reductant": unit["reductant"],
                        "hydrogenStatus": unit["hydrogenStatus"],
                    }
                    for unit in related_dri
                ],
            }
        )
    return sorted(assets, key=lambda asset: asset["id"])


def load_industrial_demand_assumptions(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("schemaVersion") != "1.0" or not isinstance(payload.get("methods"), dict):
        raise ValueError("industrial demand assumptions require schema version 1.0 and methods")
    required = {
        "hydrogen-electrolyser-grid-v1",
        "steel-eaf-electricity-v1",
        "cement-electricity-v1",
    }
    if not required <= set(payload["methods"]):
        raise ValueError("industrial demand assumptions are missing required methods")
    return payload


def _ordered_range(values: Mapping[str, Any]) -> tuple[float, float, float]:
    result = tuple(float(values[key]) for key in ("low", "central", "high"))
    if not 0 <= result[0] <= result[1] <= result[2]:
        raise ValueError("industrial demand assumption ranges must be non-negative and ordered")
    return result


def hydrogen_annual_demand_gwh(
    *,
    capacity_mwel: float | None,
    grid_connection_type: str | None,
    assumptions: Mapping[str, Any],
) -> dict[str, float] | None:
    if capacity_mwel is None or capacity_mwel < 0:
        return None
    method = assumptions["methods"]["hydrogen-electrolyser-grid-v1"]
    shares = method["gridShareByConnection"].get(grid_connection_type)
    if shares is None:
        return None
    utilization = _ordered_range(method["capacityFactor"])
    grid_share = _ordered_range(shares)
    return {
        key: round(capacity_mwel * 8.76 * utilization[index] * grid_share[index], 6)
        for index, key in enumerate(("low", "central", "high"))
    }


def steel_eaf_annual_demand_gwh(
    *, capacity_ttpa: float | None, assumptions: Mapping[str, Any]
) -> dict[str, float] | None:
    if capacity_ttpa is None or capacity_ttpa < 0:
        return None
    intensity = _ordered_range(
        assumptions["methods"]["steel-eaf-electricity-v1"]
        ["electricityIntensityMwhPerTonne"]
    )
    return {
        key: round(capacity_ttpa * intensity[index], 6)
        for index, key in enumerate(("low", "central", "high"))
    }


def cement_annual_demand_gwh(
    *, capacity_mtpa: float | None, assumptions: Mapping[str, Any]
) -> dict[str, float] | None:
    if capacity_mtpa is None or capacity_mtpa < 0:
        return None
    intensity = _ordered_range(
        assumptions["methods"]["cement-electricity-v1"]
        ["electricityIntensityMwhPerTonne"]
    )
    return {
        key: round(capacity_mtpa * 1_000 * intensity[index], 6)
        for index, key in enumerate(("low", "central", "high"))
    }


def apply_industrial_demand_model(
    asset: Mapping[str, Any], *, assumptions: Mapping[str, Any]
) -> dict[str, Any]:
    if asset.get("category") != "industrial_load" or not asset.get("gridDemandEligible"):
        return dict(asset)
    subtype = asset.get("subtype")
    method_id: str | None = None
    demand: dict[str, float] | None = None
    if subtype == "hydrogen_production":
        method_id = "hydrogen-electrolyser-grid-v1"
        demand = hydrogen_annual_demand_gwh(
            capacity_mwel=asset.get("reportedCapacity"),
            grid_connection_type=asset.get("gridConnectionType"),
            assumptions=assumptions,
        )
    elif subtype == "steel_plant":
        method_id = "steel-eaf-electricity-v1"
        demand = steel_eaf_annual_demand_gwh(
            capacity_ttpa=asset.get("electricArcCapacityTtpa"),
            assumptions=assumptions,
        )
    elif subtype == "cement_plant":
        method_id = "cement-electricity-v1"
        demand = cement_annual_demand_gwh(
            capacity_mtpa=asset.get("cementCapacityMtpa"),
            assumptions=assumptions,
        )
    if demand is None or method_id is None:
        return dict(asset)
    modelled = dict(asset)
    modelled.update(
        {
            "annualDemandGwh": demand,
            "demandMw": {
                key: round(float(demand[key]) / 8.76, 6)
                for key in ("low", "central", "high")
            },
            "demandMethodId": method_id,
            "gridDemandContribution": True,
            "valueKind": "estimated",
        }
    )
    return modelled


def _country_mappings(
    countries: Mapping[str, Any],
) -> tuple[dict[str, str], dict[str, str]]:
    iso3_to_iso2: dict[str, str] = {}
    name_to_iso2: dict[str, str] = {}
    for feature in countries.get("features", []):
        properties = feature.get("properties") or {}
        iso2 = str(properties.get("id") or properties.get("country") or feature.get("id") or "").strip().upper()
        iso3 = str(properties.get("iso3") or "").strip().upper()
        name = _key(properties.get("name"))
        if len(iso2) != 2:
            continue
        if len(iso3) == 3:
            iso3_to_iso2[iso3] = iso2
        if name:
            name_to_iso2[name] = iso2
    aliases = {
        "united_states_of_america": "US", "united_states": "US",
        "russia": "RU", "turkiye": "TR", "taiwan": "TW",
        "south_korea": "KR", "north_korea": "KP", "czechia": "CZ",
        "ivory_coast": "CI", "bolivia": "BO", "venezuela": "VE",
        "iran": "IR", "syria": "SY", "tanzania": "TZ", "vietnam": "VN",
        "laos": "LA", "moldova": "MD", "brunei": "BN",
    }
    name_to_iso2.update({name: code for name, code in aliases.items() if code in set(iso3_to_iso2.values())})
    return iso3_to_iso2, name_to_iso2


def _resolve_city_coordinates(
    assets: Iterable[Mapping[str, Any]], cities: Mapping[str, Any]
) -> list[dict[str, Any]]:
    city_index: dict[tuple[str, str], list[float]] = {}
    for feature in cities.get("features", []):
        properties = feature.get("properties") or {}
        geometry = feature.get("geometry") or {}
        coordinates = geometry.get("coordinates")
        key = (str(properties.get("country") or "").upper(), _key(properties.get("name")))
        if len(key[0]) == 2 and key[1] and isinstance(coordinates, list) and len(coordinates) == 2:
            city_index.setdefault(key, [float(coordinates[0]), float(coordinates[1])])
    resolved: list[dict[str, Any]] = []
    for raw in assets:
        asset = dict(raw)
        if asset.get("coordinates") is None:
            coordinates = city_index.get(
                (str(asset.get("country") or "").upper(), _key(asset.get("locationName")))
            )
            if coordinates is not None:
                asset["coordinates"] = coordinates
                asset["locationPrecision"] = "city_centroid"
        resolved.append(asset)
    return resolved


def load_industrial_assets_from_paths(
    source_paths: Mapping[str, Path],
    *,
    countries: Mapping[str, Any],
    admin1: Mapping[str, Any],
    cities: Mapping[str, Any],
    assumptions: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    iso3_to_iso2, country_name_to_iso2 = _country_mappings(countries)
    normalized: list[dict[str, Any]] = []
    production_path = source_paths.get(HYDROGEN_PRODUCTION_SOURCE_ID)
    if production_path is not None:
        normalized.extend(
            parse_hydrogen_production(production_path, iso3_to_iso2=iso3_to_iso2)
        )
    infrastructure_path = source_paths.get(HYDROGEN_INFRASTRUCTURE_SOURCE_ID)
    if infrastructure_path is not None:
        normalized.extend(
            parse_hydrogen_infrastructure(
                infrastructure_path, iso3_to_iso2=iso3_to_iso2
            )
        )
    cement_path = source_paths.get(GEM_CEMENT_SOURCE_ID)
    if cement_path is not None:
        normalized.extend(
            parse_gem_cement(
                cement_path, country_name_to_iso2=country_name_to_iso2
            )
        )
    steel_paths = (
        source_paths.get(GEM_STEEL_PLANT_SOURCE_ID),
        source_paths.get(GEM_STEEL_UNIT_SOURCE_ID),
        source_paths.get(GEM_IRON_UNIT_SOURCE_ID),
    )
    if all(path is not None for path in steel_paths):
        normalized.extend(
            parse_gem_steel(
                steel_paths[0],
                steel_units_path=steel_paths[1],
                iron_units_path=steel_paths[2],
                country_name_to_iso2=country_name_to_iso2,
            )
        )
    modelled = [
        apply_industrial_demand_model(asset, assumptions=assumptions)
        for asset in _resolve_city_coordinates(normalized, cities)
    ]
    by_country: dict[str, list[dict[str, Any]]] = {}
    for feature in admin1.get("features", []):
        country = str((feature.get("properties") or {}).get("country") or "").upper()
        if country:
            by_country.setdefault(country, []).append(feature)
    indexes = {
        country: build_geography_index(features)
        for country, features in by_country.items()
    }
    mappable: list[dict[str, Any]] = []
    for asset in modelled:
        if asset.get("coordinates") is None:
            continue
        country = str(asset.get("country") or "").upper()
        asset["geographyId"] = country
        assignment_asset = asset
        if asset.get("locationPrecision") != "exact":
            assignment_asset = {**asset, "locationPrecision": "exact"}
        geography_id = assign_asset_geography(
            assignment_asset, indexes.get(country, [])
        )
        asset["geographyId"] = geography_id
        if geography_id != country:
            asset["admin1Id"] = geography_id
        mappable.append(asset)
    forecast_eligible = sum(
        bool(asset.get("gridDemandContribution"))
        and asset.get("lifecycle") in {
            "announced", "planning_filed", "permitted", "pre_construction", "under_construction"
        }
        and isinstance(asset.get("targetYear"), int)
        and 2026 <= asset["targetYear"] <= 2031
        and asset.get("geographyId") != asset.get("country")
        for asset in mappable
    )
    return sorted(mappable, key=lambda asset: asset["id"]), {
        "normalized": len(normalized),
        "mappable": len(mappable),
        "forecastEligible": forecast_eligible,
    }


def merge_industrial_assets(
    registry: Mapping[str, Any], assets: Iterable[Mapping[str, Any]]
) -> dict[str, Any]:
    merged = dict(registry)
    by_id = {
        str(asset.get("id")): dict(asset)
        for asset in registry.get("assets", [])
        if asset.get("id")
    }
    for asset in assets:
        identifier = str(asset.get("id") or "").strip()
        if not identifier:
            raise ValueError("industrial assets require stable IDs")
        by_id[identifier] = dict(asset)
    merged["assets"] = [by_id[identifier] for identifier in sorted(by_id)]
    return merged
