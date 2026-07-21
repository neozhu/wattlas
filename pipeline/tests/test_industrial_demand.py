from pathlib import Path

from openpyxl import Workbook

from grid_scope.industrial_demand import (
    apply_industrial_demand_model,
    cement_annual_demand_gwh,
    hydrogen_annual_demand_gwh,
    load_industrial_demand_assumptions,
    parse_gem_cement,
    parse_gem_steel,
    parse_hydrogen_infrastructure,
    parse_hydrogen_production,
    steel_eaf_annual_demand_gwh,
)


ISO3_TO_ISO2 = {"DEU": "DE", "NLD": "NL", "AUS": "AU"}
COUNTRY_TO_ISO2 = {"germany": "DE", "namibia": "NA", "australia": "AU"}
ASSUMPTIONS_PATH = Path(__file__).resolve().parents[2] / "data" / "curated" / "industrial-demand-assumptions.json"


def _production_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hydrogen production projects"
    sheet.append(["Technology"])
    sheet.append([
        "Reference", "Project name", "Country\n(ISO-3)", "Date online",
        "Decomission date", "Status", "Technology", "Technology details",
        "Type of electricity (for electrolysis projects)",
        "If dedicated renewables, type of renewable", "Product",
        "Announced Size", "Capacity (MWel)", "Capacity (Nm³ H₂/h)",
        "Capacity\n(kt H2/y)", "Location", "Latitude", "Longitude",
    ])
    sheet.append([
        "H2-1", "Alpha Hydrogen", "DEU", 2029, None, "Feasibility study",
        "PEM", "PEM electrolysis", "Grid+Renewables", None, "H2", "80 MW", 80,
        None, None, "Bavaria", 48.1, 11.6,
    ])
    sheet.append([
        "H2-2", "Wind Hydrogen", "NLD", "2030", None, "Concept",
        "Other Electrolysis", "Alkaline", "Dedicated renewable", "Offshore wind",
        "H2", "100 MW", 100, None, None, "Rotterdam", 51.9, 4.5,
    ])
    sheet.append([
        "H2-3", "Unknown-size Hydrogen", "AUS", "2028", None, "FID/Construction",
        "ALK", "Alkaline", "Grid", None, "H2", "TBC", None, None, 20,
        "Gladstone", -23.8, 151.25,
    ])
    workbook.save(path)


def _infrastructure_workbook(path: Path) -> None:
    workbook = Workbook()
    pipeline = workbook.active
    pipeline.title = "H2 TRANSMISSION (pipe)"
    pipeline.append([
        "Ref", "Project name", "Country_1 (ISO)", "Country_2 (ISO)", "Partners",
        "Announced project date", "Announced start date", "Construction start date",
        "Date online", "Decomission date", "Repurposed_new", "Status",
        "Pipeline_type", "Announced Size", "Capacity_MWel", "Capacity_nm³ H₂/h",
        "Capacity_kt H2/y", "Length_km",
    ])
    pipeline.append([
        "PiP-001", "Alpha Backbone", "DEU", "NLD", "GridCo", None, 2028, None,
        2028, None, "New", "Feasibility study", "Onshore", "10 GW", 10_000,
        None, None, 420,
    ])

    blending = workbook.create_sheet("H2 BLENDING")
    blending.append([
        "Ref", "Project name", "Country (ISO)", "Partners", "Announced project date",
        "Announced start date", "Date online", "Decomission date", "Status",
        "H2_blending (vol)", "Announced Size", "Capacity_MWel", "Capacity_nm³ H₂/h",
        "Capacity_kt H2/y", "Length_km", "Pressure_bar", "Investment costs_MUSD",
        "Location",
    ])
    blending.append([
        "BLE-1", "Community Blend", "AUS", "GasCo", None, 2027, 2027, None,
        "FID/Construction", 0.1, "1.25 MW electrolyser", 1.25, None, None, None,
        None, 15, "Tonsley",
    ])
    workbook.save(path)


def test_hydrogen_production_normalizes_status_location_and_grid_evidence(tmp_path: Path) -> None:
    path = tmp_path / "production.xlsx"
    _production_workbook(path)

    assets = parse_hydrogen_production(path, iso3_to_iso2=ISO3_TO_ISO2)
    by_id = {asset["id"]: asset for asset in assets}

    alpha = by_id["iea-h2-production-h2-1"]
    assert alpha["country"] == "DE"
    assert alpha["lifecycle"] == "pre_construction"
    assert alpha["targetYear"] == 2029
    assert alpha["coordinates"] == [11.6, 48.1]
    assert alpha["reportedCapacity"] == 80
    assert alpha["reportedCapacityUnit"] == "MWel"
    assert alpha["gridConnectionType"] == "grid_plus_renewables"
    assert alpha["gridDemandEligible"] is True
    assert alpha["sourceRecordIds"] == ["H2-1"]

    dedicated = by_id["iea-h2-production-h2-2"]
    assert dedicated["gridConnectionType"] == "dedicated_renewable"
    assert dedicated["gridDemandEligible"] is False

    missing = by_id["iea-h2-production-h2-3"]
    assert missing["lifecycle"] == "under_construction"
    assert missing["reportedCapacity"] is None
    assert missing["gridDemandEligible"] is False
    assert missing["coordinates"] == [151.25, -23.8]


def test_hydrogen_infrastructure_is_context_only_even_with_mwel(tmp_path: Path) -> None:
    path = tmp_path / "infrastructure.xlsx"
    _infrastructure_workbook(path)

    assets = parse_hydrogen_infrastructure(path, iso3_to_iso2=ISO3_TO_ISO2)
    by_id = {asset["id"]: asset for asset in assets}

    pipeline = by_id["iea-h2-network-pip-001"]
    assert pipeline["category"] == "hydrogen_infrastructure"
    assert pipeline["subtype"] == "hydrogen_pipeline"
    assert pipeline["country"] == "DE"
    assert pipeline["secondaryCountries"] == ["NL"]
    assert pipeline["reportedCapacity"] == 10_000
    assert pipeline["reportedCapacityUnit"] == "MWel"
    assert pipeline["annualDemandGwh"] is None
    assert pipeline["gridDemandContribution"] is False

    blend = by_id["iea-h2-network-ble-1"]
    assert blend["subtype"] == "hydrogen_blending"
    assert blend["locationName"] == "Tonsley"
    assert blend["annualDemandGwh"] is None
    assert blend["gridDemandContribution"] is False


def _cement_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plant Data"
    sheet.append([
        "GEM Plant ID", "GEM Asset name (English)", "Coordinates", "Coordinate accuracy",
        "GEM wiki page", "Municipality", "Subnational unit", "Country/Area",
        "Cement Capacity (millions metric tonnes per annum)",
        "Clinker Capacity (millions metric tonnes per annum)", "Majority Cement Type",
        "Operating status", "Start date", "Owner name (English)", "Plant type",
        "Production type",
    ])
    sheet.append([
        "P-CEM-1", "Alpha Cement Plant", "-23.8, 151.25", "exact",
        "https://www.gem.wiki/Alpha_Cement_Plant", "Gladstone", "Queensland", "Australia",
        2.4, 1.8, "opc", "construction", 2028, "Alpha Cement [100%]", "integrated",
        "clinker and cement",
    ])
    workbook.save(path)


def _steel_workbooks(plant_path: Path, steel_path: Path, iron_path: Path) -> None:
    plant_workbook = Workbook()
    plant = plant_workbook.active
    plant.title = "Plant data"
    plant.append([
        "GEM plant ID", "Plant name (English)", "Owner", "Municipality",
        "Subnational unit", "Country/area", "Coordinates", "Coordinate accuracy",
        "GEM wiki page", "Power source",
    ])
    plant.append([
        "P-STEEL-1", "Alpha Green Steel", "Alpha Steel", "Bremen", "Bremen", "Germany",
        "53.1, 8.8", "exact", "https://www.gem.wiki/Alpha_Green_Steel", "grid",
    ])
    capacity = plant_workbook.create_sheet("Plant capacities and status")
    capacity.append([
        "GEM plant ID", "Plant name (English)", "Country/area", "Main production equipment",
        "Status", "Start date", "Nominal crude steel capacity (ttpa)",
        "Nominal BOF steel capacity (ttpa)", "Nominal EAF steel capacity (ttpa)",
        "Nominal IF steel capacity (ttpa)", "Nominal DRI capacity (ttpa)",
        "Nominal BF capacity (ttpa)",
    ])
    capacity.append([
        "P-STEEL-1", "Alpha Green Steel", "Germany", "EAF, DRI", "construction", 2029,
        2200, None, 1200, None, 1000, None,
    ])
    plant_workbook.save(plant_path)

    steel_workbook = Workbook()
    eaf = steel_workbook.active
    eaf.title = "Electric arc furnaces"
    eaf.append([
        "GEM plant ID", "GEM unit ID", "Unit name", "GEM wiki page", "Country/area",
        "Unit status", "Announced date", "Construction date", "Start date",
        "Current capacity (ttpa)",
    ])
    eaf.append([
        "P-STEEL-1", "U-EAF-1", "EAF 1", "https://www.gem.wiki/Alpha_Green_Steel",
        "Germany", "construction", 2025, 2027, 2029, 1200,
    ])
    steel_workbook.save(steel_path)

    iron_workbook = Workbook()
    dri = iron_workbook.active
    dri.title = "Direct reduced iron furnaces"
    dri.append([
        "GEM plant ID", "GEM unit ID", "Unit name", "GEM wiki page", "Country/area",
        "Unit status", "Announced date", "Construction date", "Start date", "Furnace type",
        "Current capacity (ttpa)", "Current or initial Reductant", "Hydrogen reductant status",
    ])
    dri.append([
        "P-STEEL-1", "U-DRI-1", "DRI 1", "https://www.gem.wiki/Alpha_Green_Steel",
        "Germany", "announced", 2025, None, 2029, "shaft furnace", 1000,
        "natural gas", "hydrogen-ready",
    ])
    iron_workbook.save(iron_path)


def test_gem_cement_preserves_capacity_status_and_project_page(tmp_path: Path) -> None:
    path = tmp_path / "cement.xlsx"
    _cement_workbook(path)

    assets = parse_gem_cement(path, country_name_to_iso2=COUNTRY_TO_ISO2)

    assert len(assets) == 1
    asset = assets[0]
    assert asset["id"] == "gem-cement-p-cem-1"
    assert asset["country"] == "AU"
    assert asset["coordinates"] == [151.25, -23.8]
    assert asset["lifecycle"] == "under_construction"
    assert asset["targetYear"] == 2028
    assert asset["cementCapacityMtpa"] == 2.4
    assert asset["clinkerCapacityMtpa"] == 1.8
    assert asset["projectUrl"] == "https://www.gem.wiki/Alpha_Cement_Plant"


def test_gem_steel_joins_plant_eaf_and_dri_evidence(tmp_path: Path) -> None:
    plant_path = tmp_path / "plants.xlsx"
    steel_path = tmp_path / "steel-units.xlsx"
    iron_path = tmp_path / "iron-units.xlsx"
    _steel_workbooks(plant_path, steel_path, iron_path)

    assets = parse_gem_steel(
        plant_path,
        steel_units_path=steel_path,
        iron_units_path=iron_path,
        country_name_to_iso2=COUNTRY_TO_ISO2,
    )

    assert len(assets) == 1
    asset = assets[0]
    assert asset["id"] == "gem-steel-p-steel-1-construction-2029"
    assert asset["country"] == "DE"
    assert asset["coordinates"] == [8.8, 53.1]
    assert asset["lifecycle"] == "under_construction"
    assert asset["targetYear"] == 2029
    assert asset["electricArcCapacityTtpa"] == 1200
    assert asset["driCapacityTtpa"] == 1000
    assert asset["sourceRecordIds"] == ["P-STEEL-1", "U-DRI-1", "U-EAF-1"]
    assert asset["sourceIds"] == [
        "gem-global-iron-units-2026",
        "gem-global-steel-plants-2026",
        "gem-global-steel-units-2026",
    ]
    assert asset["projectUrl"] == "https://www.gem.wiki/Alpha_Green_Steel"


def test_hydrogen_capacity_uses_utilization_and_grid_share_ranges() -> None:
    assumptions = load_industrial_demand_assumptions(ASSUMPTIONS_PATH)

    grid = hydrogen_annual_demand_gwh(
        capacity_mwel=100, grid_connection_type="grid", assumptions=assumptions
    )
    mixed = hydrogen_annual_demand_gwh(
        capacity_mwel=100,
        grid_connection_type="grid_plus_renewables",
        assumptions=assumptions,
    )

    assert grid == {"low": 275.94, "central": 474.354, "high": 788.4}
    assert mixed == {"low": 76.65, "central": 249.66, "high": 591.3}
    assert hydrogen_annual_demand_gwh(
        capacity_mwel=100,
        grid_connection_type="dedicated_renewable",
        assumptions=assumptions,
    ) is None


def test_steel_and_cement_intensity_conversions_preserve_units() -> None:
    assumptions = load_industrial_demand_assumptions(ASSUMPTIONS_PATH)

    assert steel_eaf_annual_demand_gwh(
        capacity_ttpa=1200, assumptions=assumptions
    ) == {"low": 360.0, "central": 523.2, "high": 924.0}
    assert cement_annual_demand_gwh(
        capacity_mtpa=2.4, assumptions=assumptions
    ) == {"low": 216.0, "central": 266.4, "high": 360.0}


def test_demand_model_keeps_missing_and_context_only_inputs_unavailable() -> None:
    assumptions = load_industrial_demand_assumptions(ASSUMPTIONS_PATH)
    context = {
        "category": "hydrogen_infrastructure", "subtype": "hydrogen_pipeline",
        "reportedCapacity": 10_000, "gridDemandEligible": False,
        "gridDemandContribution": False, "annualDemandGwh": None,
    }
    missing = {
        "category": "industrial_load", "subtype": "hydrogen_production",
        "reportedCapacity": None, "gridConnectionType": "grid", "gridDemandEligible": False,
        "gridDemandContribution": False, "annualDemandGwh": None,
    }

    assert apply_industrial_demand_model(context, assumptions=assumptions) == context
    assert apply_industrial_demand_model(missing, assumptions=assumptions) == missing


def test_demand_model_attaches_method_and_estimated_range() -> None:
    assumptions = load_industrial_demand_assumptions(ASSUMPTIONS_PATH)
    project = {
        "category": "industrial_load", "subtype": "steel_plant",
        "electricArcCapacityTtpa": 1200, "gridDemandEligible": True,
        "gridDemandContribution": False, "annualDemandGwh": None,
        "valueKind": "reported",
    }

    modelled = apply_industrial_demand_model(project, assumptions=assumptions)

    assert modelled["annualDemandGwh"] == {"low": 360.0, "central": 523.2, "high": 924.0}
    assert modelled["demandMethodId"] == "steel-eaf-electricity-v1"
    assert modelled["gridDemandContribution"] is True
    assert modelled["valueKind"] == "estimated"
